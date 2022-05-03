from selenium import webdriver
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.common.keys import Keys
import openpyxl
import sys


def check_HUD_results(browser):
    try:
        search_results = browser.find_element_by_id('ctl00_lblTotalNoRecords').text
        return search_results
    except NoSuchElementException:
        return "No properties found..."


def check_next_button_exists(browser):
    try:
        element = browser.find_element_by_name("getMoreData").submit()
        return True
    except NoSuchElementException:
        return False


def check_clasificados_results(browser):
    try:
        browser.find_element_by_tag_name('html').send_keys(Keys.PAGE_DOWN)
        search_results = browser.find_element_by_class_name("Tahoma16BrownNound").text.split('de ')[1]
        return search_results
    except NoSuchElementException:
        return "No apartments found..."


def check_deshow_results(browser):
    try:
        search_results = browser.find_element_by_class_name("page-title").text
        return search_results
    except NoSuchElementException:
        return "No apartments found..."


county = 'Guaynabo'
max_price = '150,000'  # Change code to work with max price!!!
# Also add the Homesteps Freddie Mac page!!!

workbook = openpyxl.Workbook()
sheet = workbook.active

options = webdriver.ChromeOptions()
options.add_argument('start-maximized')
# options.add_argument('ignore-certificate-errors')
# options.add_argument('incognito')
options.headless = True
browser = webdriver.Chrome(executable_path=r'./chromedriver', options=options)
browser.implicitly_wait(10)  # seconds
print('Opening FHA Seeker...')

browser.get('https://www.hudhomestore.com/Home/Index.aspx?sLanguage=ENGLISH')
state_dropdown = Select(browser.find_element_by_id('ctl00_ddState'))
state_dropdown.select_by_visible_text('PR')

browser.find_element_by_class_name('CorpFormButton').click()
# listings_found = browser.find_element_by_id('ctl00_lblTotalNoRecords').text
listings_found = check_HUD_results(browser)
sheet.title = "HUD Homestore"
sheet.column_dimensions['A'].width = 35
sheet['A1'].value = 'Number of listings found in PR'
sheet.column_dimensions['B'].width = 25
sheet['B1'].value = 'Link'
sheet['A2'].value = listings_found
sheet['B2'].value = browser.current_url
print('In progress... 25% completed')

browser.get('https://entp.hud.gov/idapp/html/condlook.cfm')
Select(browser.find_element_by_id('l_state')).select_by_visible_text('Puerto Rico')
county_field = browser.find_element_by_id('l_county')
# county_field = browser.find_element_by_id('l_city')
county_field.send_keys(county)
Select(browser.find_element_by_id('l_status_code')).select_by_visible_text('Approved')
county_field.submit()
listings_found = browser.find_element_by_class_name('textnormal').text
print('Got FHA results! Number of FHA approved condos: ' + listings_found)
FHA_approved_condos = []
next_button_exists = True
while next_button_exists:
    fonts_names = browser.find_elements_by_css_selector('a > font')
    for fonts in fonts_names:
        if fonts.text != "Exists":
            print(fonts.text)
            FHA_approved_condos.append(fonts.text)
    next_button_exists = check_next_button_exists(browser)

workbook.active = 1
workbook.create_sheet('Clasificados Online')
sheet = workbook.active
sheet.column_dimensions['A'].width = 35
sheet['A1'].value = 'Condominium Name'
sheet.column_dimensions['B'].width = 35
sheet['B1'].value = 'Number of Apartments Found'
sheet.column_dimensions['C'].width = 30
sheet['C1'].value = 'Link'

workbook.active = 2
workbook.create_sheet('deShow')
sheet = workbook.active
sheet.column_dimensions['A'].width = 35
sheet['A1'].value = 'Condominium Name'
sheet.column_dimensions['B'].width = 35
sheet['B1'].value = 'Number of Apartments Found'
sheet.column_dimensions['C'].width = 30
sheet['C1'].value = 'Link'

row_counter = 2

for condo in FHA_approved_condos:
    browser.get('https://www.clasificadosonline.com/RealEstate.asp')
    Select(browser.find_element_by_id('RESPueblos')).select_by_visible_text(county)
    text_area = browser.find_element_by_id('txtArea')
    text_area.send_keys(condo)
    browser.find_element_by_id('BtnSearchListing').click()
    workbook.active = 1
    sheet = workbook.active
    search_result = check_clasificados_results(browser)
    sheet.cell(row=row_counter, column=1).value = condo
    sheet.cell(row=row_counter, column=2).value = search_result
    sheet.cell(row=row_counter, column=3).value = browser.current_url

    # browser.get('https://deshow.com/advance-search/?operation=en-venta&type=residencial&subtipo=all&location=all'
    #             '&status=all&keyword=&price_range_min=0&price_range_max=3000000&bathrooms=all&bedrooms=all&pageid'
    #             '=1398')

    # if row_counter == 2:  # click on accept cookies button the first time only
    #     browser.find_element_by_id('btn-accept').click()
    #
    # browser.find_element_by_id('location_chosen').click()
    # browser.find_element_by_css_selector('input:focus').send_keys(county)
    # text_area = browser.find_element_by_id('keyword')
    # text_area.send_keys(condo)
    # text_area.submit()
    # workbook.active = 2
    # sheet = workbook.active
    # search_result = check_deshow_results(browser)
    # sheet.cell(row=row_counter, column=1).value = condo
    # sheet.cell(row=row_counter, column=2).value = search_result
    # sheet.cell(row=row_counter, column=3).value = browser.current_url

    row_counter += 1

print('In progress... 75% completed')
browser.get('https://www.homepath.com/')
county_field = browser.find_element_by_id('searchInput')
county_field.send_keys(county)
county_field.submit()
listings_found = browser.find_element_by_tag_name('h6').text.replace('\n', '').split('of ')[1]
browser.find_element_by_tag_name('canvas')
browser.save_screenshot(county + "_Homepath.png");

workbook.active = 3
workbook.create_sheet('Homepath')
sheet = workbook.active
sheet.column_dimensions['A'].width = 35
sheet['A1'].value = 'Number of listings found in ' + county
sheet.column_dimensions['B'].width = 30
sheet['B1'].value = 'Link'
sheet['A2'].value = listings_found
sheet['B2'].value = browser.current_url
img = openpyxl.drawing.image.Image(county + "_Homepath.png")
img.anchor = 'A3'
sheet.add_image(img)

workbook.active = 0
workbook.save(county + "_FHA_Condos.xlsx")

browser.quit()
print('Finished! Closing FHA Seeker...')
